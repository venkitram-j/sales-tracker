import logging

import openpyxl

logger = logging.getLogger("apps.core.utils")


class ExcelParseError(Exception):
    """Raised when an uploaded workbook cannot be parsed at all."""


def iter_excel_rows(file_obj, header_row=1, data_start_row=None, start_col=1):
    """Yield (row_number, {header: value}) for every non-empty data row of the first sheet.

    Supports spreadsheets that don't start their real data in row 1 /
    column A - e.g. files with a title banner, notes, or an extra index
    column before the actual table:

    - `header_row`: 1-indexed row number containing the column headers.
    - `data_start_row`: 1-indexed row number where data begins. Defaults
      to the row immediately after `header_row`.
    - `start_col`: 1-indexed column number where headers/data begin
      (1 = column A). Any columns before this are ignored entirely.

    Header cell values are normalised to lowercase, stripped strings so
    mapping is forgiving of spacing / casing differences in user-supplied
    spreadsheets.
    """
    try:
        workbook = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 - surface any parse failure uniformly
        logger.exception("Failed to open uploaded workbook")
        raise ExcelParseError(f"Could not read the uploaded file: {exc}") from exc

    sheet = workbook.active

    header_cells = next(
        sheet.iter_rows(min_row=header_row, max_row=header_row, min_col=start_col, values_only=True), None
    )
    if header_cells is None:
        workbook.close()
        raise ExcelParseError(f"Header row {header_row} was not found in the uploaded file.")

    headers = [str(h).strip().lower() if h is not None else "" for h in header_cells]
    if not any(headers):
        workbook.close()
        raise ExcelParseError(f"Header row {header_row} appears to be empty. Check the Header Row setting.")

    data_start = data_start_row or (header_row + 1)

    for row_number, raw_row in enumerate(
        sheet.iter_rows(min_row=data_start, min_col=start_col, values_only=True), start=data_start
    ):
        if raw_row is None or all(cell in (None, "") for cell in raw_row):
            continue
        row_dict = dict(zip(headers, raw_row))
        yield row_number, row_dict

    workbook.close()


def scan_text_before_row(file_obj, before_row):
    """Returns every non-empty cell value (stringified) from rows 1..before_row-1
    of the first sheet, across all columns.

    Useful for pulling metadata (report titles, period banners, notes) that
    sits above the real header row of an uploaded spreadsheet. Resets the
    file pointer to the start when done so the caller can re-read the file
    afterwards (e.g. via `iter_excel_rows`).
    """
    if before_row <= 1:
        return []
    try:
        workbook = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        logger.exception("Failed to open uploaded workbook while scanning for pre-header text")
        raise ExcelParseError(f"Could not read the uploaded file: {exc}") from exc

    sheet = workbook.active
    texts = []
    for row in sheet.iter_rows(min_row=1, max_row=before_row - 1, values_only=True):
        for cell in row:
            if cell not in (None, ""):
                texts.append(str(cell))
    workbook.close()

    if hasattr(file_obj, "seek"):
        file_obj.seek(0)

    return texts
